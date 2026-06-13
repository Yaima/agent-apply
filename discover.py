#!/usr/bin/env python3
"""Find current openings at named companies and append them to the tracker.
Usage: python discover.py --tracker YourTracker.xlsx --companies "roblox, figma, gusto, google"
Resolves each company's source automatically (Greenhouse/Ashby/Lever boards, Google, Apple,
Workday); filters by profile.search (or LLM relevance with --match llm); dedupes; appends rows.
Bot-gated sites (e.g. Salesforce) are reported as manual — add those by URL with --urls.
"""
import argparse, asyncio, json, re
import httpx, yaml
from pathlib import Path
from datetime import date
from urllib.parse import quote
from openpyxl import load_workbook
from agent.tracker import backup
from agent.ats import RECAPTCHA_V3

KNOWN = {  # proven board tokens; anything else is auto-slugged and probed
 "greenhouse": ["figma","anthropic","gusto","mercury","gleanwork","stripe","coinbase","databricks","lyft",
   "discord","affirm","roblox","pinterest","vercel","sonyinteractiveentertainmentglobal","brex","duolingo",
   "instacart","gitlab"],
 "ashby": ["openai","ramp","linear","notion","elevenlabs","sierra","perplexity"],
}
ALIAS = {"sony":"sonyinteractiveentertainmentglobal","playstation":"sonyinteractiveentertainmentglobal","glean":"gleanwork"}

# Company-owned career sites with no public board API. Some serve real HTML we can read
# with httpx (Apple); others do ALL search/filter/pagination client-side via a bot-guarded
# XHR and render the same unfiltered list to any non-interactive request — those can't be
# discovered without forging anti-bot tokens, which this tool refuses to do (see JS_GATED).
# Each fetcher returns (ats_label, [(title, location, url, salary), ...]) — the same tuple
# shape probe() returns, so dedup/filter/append downstream is unchanged.
APPLE_BASE = "https://jobs.apple.com"
APPLE_LOC = {"united states":"united-states-USA","usa":"united-states-USA","us":"united-states-USA"}
GOOGLE_BASE = "https://www.google.com/about/careers/applications/jobs/results/"
# Known Workday endpoints (host, site) for common companies — checked before auto-resolve.
WORKDAY = {
    "nvidia": ("nvidia.wd5.myworkdayjobs.com", "NVIDIAExternalCareerSite"),
    "cisco":  ("cisco.wd5.myworkdayjobs.com", "Cisco_Careers"),
}
# Datacenters and site-name templates tried when auto-resolving a Workday tenant by name.
WD_DCS = ("wd1", "wd3", "wd5", "wd103", "wd12", "wd101")
def _wd_sites(tenant):
    cap = tenant.capitalize()
    return [f"{tenant.upper()}ExternalCareerSite", f"{cap}ExternalCareerSite", "ExternalCareerSite",
            "External", f"{cap}_Careers", f"{cap}Careers", "Careers", "careers", tenant, f"{cap}External"]

# Companies whose careers sites filter only via a bot-guarded JS API: search/team/location
# and pagination params are all ignored for httpx AND a real automated browser (verified for
# Salesforce 2026-06: every request returns the full unfiltered list). Route to --urls.
JS_GATED = {"salesforce": "careers.salesforce.com filters only via a bot-guarded JS API"}

async def fetch_apple(client, query, location, max_pages):
    """jobs.apple.com — server-rendered cards; location honored via URL slug."""
    slug = APPLE_LOC.get((location or "").strip().lower())
    if location and not slug:  # weak fallback for unmapped locations
        slug = re.sub(r"[^a-z0-9]+", "-", location.lower()).strip("-") + "-USA"
    jobs = []
    for page in range(1, max_pages + 1):
        loc_q = f"&location={slug}" if slug else ""
        url = f"{APPLE_BASE}/en-us/search?search={quote(query)}{loc_q}&page={page}"
        r = await client.get(url, timeout=20)
        if r.status_code != 200: break
        html = r.text
        cards = list(re.finditer(
            r'<a class="link-inline[^"]*"[^>]*href="(/en-us/details/(\d+-\d+)/[^"?]*)[^"]*"[^>]*>([^<]+)</a>',
            html))
        if not cards: break
        for m in cards:
            href, _jid, title = m.groups()
            tail = html[m.end():m.end() + 1400]
            loc = (re.search(r'search-store-name-container[^"]*"[^>]*>([^<]+)<', tail) or [None, ""])[1].strip()
            jobs.append((title.strip(), loc, APPLE_BASE + href, ""))
    return "apple", jobs

async def fetch_google(client, query, location, max_pages):
    """google.com/about/careers — server-rendered, query-filtered. Parse per card so each
    title/location binds to its own job id (no positional zip that can silently misalign)."""
    jobs = []
    for page in range(1, max_pages + 1):
        loc_q = f"&location={quote(location)}" if location else ""
        url = f"{GOOGLE_BASE}?q={quote(query)}{loc_q}&page={page}"
        r = await client.get(url, timeout=20)
        if r.status_code != 200: break
        html = r.text
        slug_by_id = {}
        for jid, slug in re.findall(r'jobs/results/(\d+)-([a-z0-9\-]+)', html):
            slug_by_id.setdefault(jid, f"{jid}-{slug}")
        cards = re.split(r'<li class="lLd3Je"', html)[1:]
        if not cards: break
        for card in cards:
            tm = re.search(r'<h3 class="QJPWVe">([^<]+)</h3>', card)
            im = re.search(r"ssk='1\d:(\d+)'", card) or re.search(r'jobs/results/(\d+)-', card)
            if not tm or not im: continue
            full = slug_by_id.get(im.group(1))
            if not full: continue
            loc = re.sub(r"^[;\s]+", "", (re.search(r'class="r0wTof[^"]*">([^<]+)<', card) or [None, ""])[1]).strip()
            jobs.append((tm.group(1).strip(), loc, f"{GOOGLE_BASE}{full}", ""))
    return "google", jobs

async def fetch_workday(client, host, site, tenant, query, location, max_pages):
    """Workday public cxs JSON API — paginate by offset until total."""
    jobs, total = [], None
    for page in range(max_pages):
        body = {"appliedFacets": {}, "limit": 20, "offset": page * 20, "searchText": query or ""}
        r = await client.post(f"https://{host}/wday/cxs/{tenant}/{site}/jobs",
                              json=body, headers={"Content-Type": "application/json"}, timeout=20)
        if r.status_code != 200: break
        data = r.json()
        total = data.get("total") if total is None else total
        posts = data.get("jobPostings") or []
        if not posts: break
        for j in posts:
            jobs.append((j.get("title", "").strip(), (j.get("locationsText") or "").strip(),
                         f"https://{host}/{site}{j.get('externalPath','')}", ""))
        if total is not None and (page + 1) * 20 >= total: break
    return "workday", jobs

async def resolve_workday(client, tenant):
    """Best-effort: probe datacenters x site-name templates for a tenant. Returns (host, site) or None.
    Skips a whole datacenter once its host proves unreachable, to avoid a long serial request storm."""
    probe_body = {"appliedFacets": {}, "limit": 1, "offset": 0, "searchText": ""}
    for dc in WD_DCS:
        host = f"{tenant}.{dc}.myworkdayjobs.com"
        for i, site in enumerate(_wd_sites(tenant)):
            try:
                r = await client.post(f"https://{host}/wday/cxs/{tenant}/{site}/jobs",
                                      json=probe_body, headers={"Content-Type": "application/json"}, timeout=5)
            except Exception:
                if i == 0: break     # DNS/connect failure on the first try → this DC is dead, skip its templates
                continue
            if r.status_code == 200 and "total" in r.text:
                return host, site
    return None

async def fetch_from_url(client, careers_url, query, location, max_pages):
    """Reliable universal path: detect the ATS from a pasted careers URL and pull all roles."""
    from urllib.parse import urlparse
    p = urlparse(careers_url)
    host = p.netloc
    if "myworkdayjobs.com" in host:
        tenant = host.split(".")[0]
        segs = [s for s in p.path.split("/") if s and not re.fullmatch(r"[a-z]{2}-[A-Z]{2}", s)
                and s not in ("wday", "cxs", "jobs")]
        site = next((s for s in segs if s != tenant), segs[0] if segs else None)
        if site:
            return await fetch_workday(client, host, site, tenant, query, location, max_pages)
    if "jobs.apple.com" in host:
        return await fetch_apple(client, query, location, max_pages)
    if "google.com" in host and "careers" in p.path:
        return await fetch_google(client, query, location, max_pages)
    for ats, tok_rx in (("greenhouse", r"greenhouse\.io/(?:embed/job_board\?for=|[^/]*/?)?([a-z0-9]+)"),
                        ("lever", r"jobs\.lever\.co/([a-z0-9\-]+)"),
                        ("ashby", r"ashbyhq\.com/([a-z0-9\-]+)")):
        m = re.search(tok_rx, careers_url, re.I)
        if m:
            return await probe(client, m.group(1))
    return None, []

CUSTOM = {"apple": fetch_apple, "google": fetch_google}

async def probe(client, name):
    slug = ALIAS.get(name, re.sub(r"[^a-z0-9]", "", name.lower()))
    candidates = [("greenhouse", s) for s in ([slug] if slug not in KNOWN["greenhouse"] else [slug])] \
               + [("ashby", slug), ("lever", slug)]
    for ats, tok in candidates:
        try:
            if ats == "greenhouse":
                r = await client.get(f"https://boards-api.greenhouse.io/v1/boards/{tok}/jobs", timeout=12)
                if r.status_code == 200 and r.json().get("jobs"):
                    return ats, [(j["title"], j.get("location",{}).get("name",""), j["absolute_url"], "") for j in r.json()["jobs"]]
            elif ats == "ashby":
                r = await client.get(f"https://api.ashbyhq.com/posting-api/job-board/{tok}?includeCompensation=true", timeout=12)
                if r.status_code == 200 and r.json().get("jobs"):
                    return ats, [(j["title"], j.get("location",""), j.get("jobUrl") or j.get("applyUrl",""),
                                  (j.get("compensation") or {}).get("compensationTierSummary") or "") for j in r.json()["jobs"]]
            else:
                r = await client.get(f"https://api.lever.co/v0/postings/{tok}?mode=json", timeout=12)
                if r.status_code == 200 and r.json():
                    return ats, [(j["text"], (j.get("categories") or {}).get("location",""), j.get("hostedUrl",""), "") for j in r.json()]
        except Exception:
            pass
    return None, []

async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tracker", required=True)
    ap.add_argument("--companies", default=None, help="comma-separated company names")
    ap.add_argument("--urls", default=None, help="comma-separated posting URLs to append directly (Google, Apple, anywhere)")
    ap.add_argument("--company-name", default=None, help="company label for --urls rows")
    ap.add_argument("--search", default=None, help="search term for company sites (Google/Apple/Workday); defaults to profile.search.roles")
    ap.add_argument("--location", default="united states", help='location scope for company-owned sites (default "united states"; "" = all locations)')
    ap.add_argument("--max-pages", type=int, default=12, help="page cap per company-owned site")
    ap.add_argument("--careers-url", default=None, help="explicit careers/ATS URL (Workday/Greenhouse/Lever/Ashby/Google/Apple) for one company; pair with --company-name")
    ap.add_argument("--match", choices=["auto", "llm", "keyword"], default="auto",
                    help="how to match roles: 'llm' ranks by résumé+profile (needs ANTHROPIC_API_KEY), 'keyword' uses profile.search regex, 'auto' picks llm when a key is present")
    ap.add_argument("--near", default=None, help='commute area to keep roles within, e.g. "Seattle area" or "<city> <zip>" (LLM match; Remote always kept)')
    ap.add_argument("--relocate", action="store_true", help="with --near: also keep strong out-of-area roles as relocation options")
    ap.add_argument("--profile", default="profile.yaml")
    args = ap.parse_args()
    from agent import llm
    try:
        with open(args.profile) as f:
            prof = yaml.safe_load(f) or {}
    except FileNotFoundError:
        raise SystemExit(f"{args.profile} not found — copy profile.example.yaml to {args.profile} and fill it in.")
    except yaml.YAMLError as e:
        raise SystemExit(f"{args.profile} isn't valid YAML:\n  {e}")
    srch = prof.get("search", {}) or {}
    if args.near is None:
        args.near = srch.get("near")        # default commute area from profile.search.near
    def rx_from(words):
        return re.compile("|".join(re.escape(w).replace(r"\ ", r"\s+") for w in words), re.I) if words else None
    if srch.get("title_match"):           # advanced: raw regex wins
        inc, exc = re.compile(srch["title_match"], re.I), re.compile(srch.get("title_exclude") or r"$^", re.I)
        lvl = None
    else:                                  # plain-language: roles / levels / exclude lists
        inc = rx_from(srch.get("roles")) or re.compile(r"$^")
        exc = rx_from(srch.get("exclude")) or re.compile(r"$^")
        lvl = rx_from(srch.get("levels"))
    out = src = args.tracker            # one tracker: read and write the same file
    if Path(src).exists():
        backup(src)
        wb = load_workbook(src)
        if "Job Tracker" not in wb.sheetnames:
            raise SystemExit(f'{src} has no "Job Tracker" sheet (found: {wb.sheetnames}). Point --tracker at the right file.')
        sh = wb["Job Tracker"]
    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        print(f"'{args.tracker}' doesn't exist yet — creating a fresh tracker.")
        wb = Workbook(); sh = wb.active; sh.title = "Job Tracker"
        sh.append(["#","Company","Role","Level","Address / Location","City/Area","Mode","Approx Mi",
                   "Caltrain<1mi","Salary","Application URL","Notes","Status","Applied?","Date",
                   "Follow-up","ATS"])
        for c in range(1, 18): sh.cell(1, c).font = Font(bold=True)
        for col, w in {"B":18,"C":46,"F":22,"J":22,"K":60,"L":26,"M":22}.items():
            sh.column_dimensions[col].width = w
    def jobid(u):
        ids = re.findall(r"(\d{6,})", str(u))
        return ids[-1] if ids else None     # last long number = the job id (not a board/segment id)
    seen_urls, seen_ids = set(), set()       # full URLs + (company, job-id) — id scoped to company
    def is_dup(company, url):
        return url in seen_urls or (jobid(url) and (str(company).strip().lower(), jobid(url)) in seen_ids)
    def remember(company, url):
        seen_urls.add(url)
        if jobid(url): seen_ids.add((str(company).strip().lower(), jobid(url)))
    for r in range(2, sh.max_row + 1):
        u = str(sh.cell(r, 11).value or "")
        if u: remember(sh.cell(r, 2).value, u)
    added = 0
    if args.urls:
        for u in [x.strip() for x in args.urls.split(",") if x.strip()]:
            label = args.company_name or re.sub(r"^www\.|\..*$", "", u.split("//")[-1].split("/")[0]).title()
            if is_dup(label, u): continue
            remember(label, u)
            r = sh.max_row + 1
            sh.cell(r,1).value = r - 1
            sh.cell(r,2).value = label
            sh.cell(r,3).value = "(added by URL — title on posting)"
            sh.cell(r,11).value = u
            sh.cell(r,12).value = f"Added by URL {date.today().isoformat()}"
            added += 1
        print(f"{added} URL row(s) appended")
    if not args.companies and not args.careers_url:
        wb.save(out); print(f"-> {out}"); return
    # one or more role queries (comma-separated --search, else profile.search.roles)
    queries = [q.strip() for q in (args.search or "").split(",") if q.strip()] or list(srch.get("roles") or [])
    use_llm = args.match == "llm" or (args.match == "auto" and llm.available())
    if args.match == "llm" and not llm.available():
        print("note: --match llm needs ANTHROPIC_API_KEY — falling back to keyword matching"); use_llm = False
    if args.near and not use_llm:
        print('note: --near filtering needs LLM matching (set ANTHROPIC_API_KEY); ignoring location for now')
    brief = llm.candidate_brief(prof) if use_llm else None
    if use_llm:
        print(f"matching: LLM relevance (résumé + profile)" + (f" within {args.near}" if args.near else ""))

    async def union(fetch_one):
        """Run a query-driven fetcher once per role query; union jobs by URL. Returns (ats, jobs)."""
        ats, seen_u, acc = None, set(), []
        for q in (queries or [""]):
            a, js = await fetch_one(q)
            ats = a or ats
            for j in js:
                if j[2] not in seen_u:
                    seen_u.add(j[2]); acc.append(j)
        return ats, acc

    def keep_decisions(jobs):
        """Return list of (keep, reason) aligned to jobs — LLM ranking (role + location) or keyword regex."""
        if use_llm:
            return llm.rank_roles([(t, loc) for t, loc, _, _ in jobs], brief, args.near or "", args.relocate)
        out = []
        for title, *_ in jobs:
            ok = bool(inc.search(title)) and not exc.search(title) and (not lvl or lvl.search(title))
            out.append((ok, ""))
        return out

    def append_jobs(name, ats, jobs):
        nonlocal added
        new = 0
        for (title, loc, url, sal), (keep, reason) in zip(jobs, keep_decisions(jobs)):
            if not keep: continue
            if is_dup(name, url): continue
            remember(name, url)
            r = sh.max_row + 1
            sh.cell(r,1).value = r - 1; sh.cell(r,2).value = name.title()
            sh.cell(r,3).value = title; sh.cell(r,6).value = loc
            sh.cell(r,10).value = sal; sh.cell(r,11).value = url
            sh.cell(r,12).value = (f"Discovered {date.today().isoformat()}"
                                   + (f" — {reason}" if reason else ""))
            sh.cell(r,17).value = (f"{ats.title()} (reCAPTCHA v3 — apply by hand)" if ats in RECAPTCHA_V3
                                   else f"{ats.title()} ({'site' if ats in CUSTOM else 'auto'})")
            new += 1; added += 1
        print(f"{name}: {len(jobs)} open roles on {ats}, {new} new matching rows appended")

    async with httpx.AsyncClient(headers={"User-Agent": "Mozilla/5.0"}) as client:
        if args.careers_url:
            ats, jobs = await union(lambda q: fetch_from_url(client, args.careers_url, q, args.location, args.max_pages))
            name = args.company_name or re.sub(r"^www\.|\..*$", "", args.careers_url.split("//")[-1].split("/")[0])
            if not jobs:
                print(f"{name}: couldn't read roles from {args.careers_url} — unsupported ATS, or pass a Workday cxs / board URL")
            else:
                append_jobs(name, ats, jobs)
        for name in [c.strip() for c in (args.companies or "").split(",") if c.strip()]:
            slug = ALIAS.get(name.lower(), re.sub(r"[^a-z0-9]", "", name.lower()))
            if slug in JS_GATED:
                print(f"{name}: {JS_GATED[slug]} — not auto-discoverable. Pick postings on the site and add them with:\n"
                      f"   python discover.py --tracker {args.tracker} --urls \"<posting url>,...\" --company-name {name.title()}")
                continue
            try:
                if slug in CUSTOM:
                    if not any(q.strip() for q in queries):
                        print(f"{name}: needs a search term — pass --search \"...\" or set profile.search.roles"); continue
                    ats, jobs = await union(lambda q: CUSTOM[slug](client, q, args.location, args.max_pages))
                else:
                    ats, jobs = await probe(client, name)
                    if not jobs:  # try Workday (registry, then best-effort auto-resolve)
                        wd = WORKDAY.get(slug) or await resolve_workday(client, slug)
                        if wd:
                            ats, jobs = await union(lambda q: fetch_workday(client, wd[0], wd[1], slug, q, args.location, args.max_pages))
            except Exception as e:
                print(f"{name}: discovery error ({type(e).__name__}: {str(e)[:100]}) — skipping"); continue
            if not jobs:
                print(f"{name}: no public board found (Greenhouse/Ashby/Lever/Workday/Google/Apple). "
                      f"If it's on another ATS, pass --careers-url \"<careers url>\" --company-name {name.title()}")
                continue
            append_jobs(name, ats, jobs)
    try:
        wb.save(out)
    except PermissionError:
        out = out.replace(".xlsx", f".rescued-{date.today().isoformat()}.xlsx")
        wb.save(out)
        print(f"(tracker was locked — open in Excel? — saved this run to {out} instead)")
    print(f"\n{added} rows appended -> {out}. Apply with: python run.py --tracker {args.tracker} --company <name>")

if __name__ == "__main__":
    asyncio.run(main())
