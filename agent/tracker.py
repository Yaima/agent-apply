import shutil
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

COLS = {"status": 13, "applied": 14, "date": 15, "followup": 16}  # M,N,O,P (1-indexed)

def backup(path, keep=10):
    """Copy the tracker into backups/ once per run, so editing it in place can't lose data.
    Keeps the newest `keep` snapshots. Returns the backup path, or None if the file is new."""
    p = Path(path)
    if not p.exists():
        return None
    bdir = p.parent / "backups"; bdir.mkdir(exist_ok=True)
    dest = bdir / f"{p.stem}.{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(p, dest)
    for old in sorted(bdir.glob(f"{p.stem}.*.xlsx"))[:-keep]:
        try: old.unlink()
        except OSError: pass
    return dest

class Tracker:
    def __init__(self, path: str, out_path: str):
        self.wb = load_workbook(path)
        self.sh = self.wb["Job Tracker"]
        self.out_path = out_path
        self.ex = self.wb["Exceptions"] if "Exceptions" in self.wb.sheetnames else self._make_ex()

    def _make_ex(self):
        ex = self.wb.create_sheet("Exceptions")
        ex.append(["Row", "Company", "Role", "URL", "Reason", "Detail", "Date"])
        return ex

    def rows_to_apply(self):
        """Yields (excel_row, company, role, url) for rows with no Applied? mark."""
        for r in range(2, self.sh.max_row + 1):
            url = self.sh.cell(r, 11).value
            applied = self.sh.cell(r, COLS["applied"]).value
            if url and not applied:
                yield r, self.sh.cell(r, 2).value, self.sh.cell(r, 3).value, str(url).strip()

    def mark_applied(self, r: int, confirmation: str):
        self.sh.cell(r, COLS["status"]).value = "Applied (auto)"
        self.sh.cell(r, COLS["applied"]).value = "Yes"
        self.sh.cell(r, COLS["date"]).value = date.today().isoformat()
        self.sh.cell(r, COLS["followup"]).value = confirmation[:120]
        self.save()

    def mark_exception(self, r, company, role, url, reason, detail=""):
        self.sh.cell(r, COLS["status"]).value = f"Exception: {reason}"
        self.ex.append([r, company, role, url, reason, str(detail)[:300], date.today().isoformat()])
        self.save()

    def mark_dryrun(self, r):
        self.sh.cell(r, COLS["status"]).value = f"Dry-run OK {date.today().isoformat()}"
        self.save()

    def mark_closed(self, r):
        self.sh.cell(r, COLS["status"]).value = "Posting closed"
        self.save()

    def mark_needs_check(self, r, detail=""):
        # the form was submitted but no confirmation was seen — mark Applied? non-empty so the
        # row is never auto-resubmitted, and flag it for the human to verify.
        self.sh.cell(r, COLS["status"]).value = "Submitted? verify by hand"
        self.sh.cell(r, COLS["applied"]).value = "Check"
        self.sh.cell(r, COLS["date"]).value = date.today().isoformat()
        self.sh.cell(r, COLS["followup"]).value = str(detail)[:120]
        self.save()

    def save(self):
        self.wb.save(self.out_path)
