"""Tracker: row selection (no double-apply), the submitted-but-unconfirmed guard, backups."""
import os, shutil, tempfile, unittest
from openpyxl import Workbook, load_workbook
from agent.tracker import Tracker, backup, COLS


def make_tracker(rows):
    """rows: list of (company, role, url, applied_mark). Returns a temp .xlsx path."""
    d = tempfile.mkdtemp()
    path = os.path.join(d, "t.xlsx")
    wb = Workbook(); sh = wb.active; sh.title = "Job Tracker"
    sh.append(["#"] + [""] * 16)
    for i, (co, role, url, applied) in enumerate(rows, 1):
        r = sh.max_row + 1
        sh.cell(r, 1).value = i; sh.cell(r, 2).value = co; sh.cell(r, 3).value = role
        sh.cell(r, 11).value = url
        if applied: sh.cell(r, COLS["applied"]).value = applied
    wb.save(path)
    return path


class TestTracker(unittest.TestCase):
    def setUp(self):
        self.path = make_tracker([
            ("Figma", "Designer", "https://x/jobs/100001", None),     # row 2: open
            ("Roblox", "PD", "https://x/jobs/100002", "Yes"),         # row 3: already applied
            ("Stripe", "Visual", "https://x/jobs/100003", None),      # row 4: open
        ])
        self.tr = Tracker(self.path, self.path)

    def tearDown(self):
        shutil.rmtree(os.path.dirname(self.path), ignore_errors=True)

    def test_rows_to_apply_skips_applied(self):
        rows = list(self.tr.rows_to_apply())
        excel_rows = [r[0] for r in rows]
        self.assertEqual(excel_rows, [2, 4])          # row 3 (Applied?=Yes) excluded

    def test_mark_applied_excludes_from_next_run(self):
        self.tr.mark_applied(2, "confirmation #abc")
        self.assertNotIn(2, [r[0] for r in self.tr.rows_to_apply()])

    def test_needs_check_prevents_double_apply(self):
        # submitted-but-unconfirmed must NOT be re-applied next run
        self.tr.mark_needs_check(4, "no confirmation seen")
        rows = [r[0] for r in self.tr.rows_to_apply()]
        self.assertNotIn(4, rows)
        sh = load_workbook(self.path)["Job Tracker"]
        self.assertEqual(sh.cell(4, COLS["applied"]).value, "Check")

    def test_mark_closed_sets_status(self):
        self.tr.mark_closed(2)
        sh = load_workbook(self.path)["Job Tracker"]
        self.assertIn("closed", str(sh.cell(2, COLS["status"]).value).lower())

    def test_exceptions_sheet_created(self):
        self.tr.mark_exception(2, "Figma", "Designer", "https://x", "login-wall", "detail")
        self.assertIn("Exceptions", load_workbook(self.path).sheetnames)


class TestBackup(unittest.TestCase):
    def test_backup_snapshots_into_backups_dir(self):
        path = make_tracker([("A", "B", "https://x/jobs/1", None)])
        dest = backup(path)
        self.assertTrue(os.path.exists(dest))
        self.assertEqual(os.path.basename(os.path.dirname(dest)), "backups")
        shutil.rmtree(os.path.dirname(path), ignore_errors=True)

    def test_backup_of_missing_file_is_noop(self):
        self.assertIsNone(backup("/tmp/definitely_not_here_12345.xlsx"))


if __name__ == "__main__":
    unittest.main()
